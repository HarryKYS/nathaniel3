import pandas
import numpy
import time

from openpyxl import load_workbook
from datetime import date, timedelta
from openpyxl.utils import get_column_letter

today = date.today()
yesterday = date.today() - timedelta(1)


harry1 = today.strftime('%Y-%m-%d')
# wb = load_workbook("CMAR_" + harry1 + ".xlsx")
# wb = load_workbook('harry1.xlsx')

# ws.auto_filter.add_filter_column(12,[f"{harry1}"+ "23:59:59"])
r_csv = pandas.read_csv("CMAR_" + f"{harry1}" + ".csv", encoding="cp949")

save_xlsx = pandas.ExcelWriter("CMAR_" + f"{harry1}" + ".xlsx")
r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환

save_xlsx.save() #xlsx 파일로 저장

time.sleep(2)

wb = load_workbook("CMAR_" + harry1 + ".xlsx")
# wb = load_workbook('harry1.xlsx')

ws = wb.active

# 컬럼 0부터 시작
# ws.auto_filter.ref = "A1:AK30000"
# ws.dimensions 전체 범위
ws.auto_filter.ref = ws.dimensions
ws.auto_filter.add_filter_column(5,['GUR_2', 'JNG_1', 'NYJ_3', 'NYJ_4', 'SDG_1'])
ws.auto_filter.add_filter_column(8,['FRESH_OVERNIGHT', 'OVERNIGHT'])
ws.auto_filter.add_filter_column(12,[f"{harry1}"+ " 07:00:00"])
row_max = ws.max_row
col_max = ws.max_column
for r in range(1,row_max):
    ws[f'C{r}'].number_format ='#'
    ws[f'AF{r}'].number_format ='#'

# ws.auto_filter.add_sort_condition("F2:F30000")
# sorted(key=lambda F:F[1])
# ws.auto_filter.add_sort_condition("F2:F30000")
for col in range(1, col_max):
    ws.column_dimensions[get_column_letter(col)].width = 17

wb.save("W1 "+f"{harry1}"+".xlsx")
